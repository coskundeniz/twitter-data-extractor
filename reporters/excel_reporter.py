from string import ascii_uppercase

import openpyxl
from openpyxl.styles import Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import IllegalCharacterError

from models.user import User
from models.tweet import Tweet
from reporters.file_reporter import FileReporter
from utils import logger, ExtractedDataType


MAX_ROWS_IN_SHEET = 1048576


class ExcelReporter(FileReporter):
    """Excel report generator

    :type filename: str
    :param filename: Name of the output file
    :type extracted_data_type: ExtractedDataType
    :param extracted_data_type: Enum value for the extracted data type
    """

    def __init__(self, filename: str, extracted_data_type: ExtractedDataType) -> None:

        super().__init__(filename, extracted_data_type)

        self._sheet = None
        self._second_sheet = None

    def _save_user_data(self, extracted_data: User) -> None:
        """Save single user data

        :type extracted_data: User
        :param extracted_data: User object
        """

        logger.info("Saving user data...")

        workbook = self._create_workbook()
        self._sheet = workbook.active

        logger.info(extracted_data)

        self._add_user_header(self._sheet)

        try:
            self._sheet.append(FileReporter._get_user_row_data(extracted_data.data))
        except IllegalCharacterError:
            extracted_data.data["description"] = ILLEGAL_CHARACTERS_RE.sub(
                r"", extracted_data.data["description"]
            )
            extracted_data.data["name"] = ILLEGAL_CHARACTERS_RE.sub(
                r"", extracted_data.data["name"]
            )
            extracted_data.data["pinned_tweet_text"] = ILLEGAL_CHARACTERS_RE.sub(
                r"", extracted_data.data["pinned_tweet_text"]
            )

            try:
                self._sheet.append(FileReporter._get_user_row_data(extracted_data.data))
            except IllegalCharacterError:
                logger.error(
                    f"Received IllegalCharacterError for {extracted_data.data['username']}. Skipping..."
                )

        self._adjust_column_widths()

        workbook.save(self._filename)

    def _save_users_data(self, extracted_data: list[User]) -> None:
        """Save users/friends/followers data

        :type extracted_data: list
        :param extracted_data: List of Users(users/friends/followers)
        """

        is_friends_data = self._extracted_data_type == ExtractedDataType.FRIENDS

        if self._extracted_data_type == ExtractedDataType.USERS:
            logger.debug("Saving users data...")
        else:
            logger.debug(f"Saving {'friends' if is_friends_data else 'followers'} data...")

        workbook = self._create_workbook()
        self._sheet = workbook.active

        self._add_user_header(self._sheet)

        rows_added = 1

        for user_data_item in extracted_data:
            try:
                if rows_added == MAX_ROWS_IN_SHEET - 1:
                    logger.debug(
                        "Max row per worksheet limit will be reached. Creating new worksheet..."
                    )
                    self._second_sheet = workbook.create_sheet()
                    self._add_user_header(self._second_sheet)

                if rows_added < MAX_ROWS_IN_SHEET:
                    self._sheet.append(FileReporter._get_user_row_data(user_data_item.data))
                else:
                    self._second_sheet.append(FileReporter._get_user_row_data(user_data_item.data))

                rows_added += 1

            except IllegalCharacterError:
                user_data_item.data["description"] = ILLEGAL_CHARACTERS_RE.sub(
                    r"", user_data_item.data["description"]
                )
                user_data_item.data["name"] = ILLEGAL_CHARACTERS_RE.sub(
                    r"", user_data_item.data["name"]
                )
                user_data_item.data["pinned_tweet_text"] = ILLEGAL_CHARACTERS_RE.sub(
                    r"", user_data_item.data["pinned_tweet_text"]
                )

                try:
                    if rows_added < MAX_ROWS_IN_SHEET:
                        self._sheet.append(FileReporter._get_user_row_data(user_data_item.data))
                    else:
                        self._second_sheet.append(
                            FileReporter._get_user_row_data(user_data_item.data)
                        )

                    rows_added += 1
                except IllegalCharacterError:
                    logger.error(
                        f"Received IllegalCharacterError for {user_data_item.data['username']}. Skipping..."
                    )

        self._adjust_column_widths()

        workbook.save(self._filename)

    def _save_tweets_data(self, extracted_data: list[Tweet]) -> None:
        """Save tweets data

        :type extracted_data: list
        :param extracted_data: List of Tweets
        """

        logger.debug("Saving tweets data...")

        workbook = self._create_workbook()
        self._sheet = workbook.active

        self._add_tweet_header(self._sheet)

        rows_added = 1

        for tweet_data_item in extracted_data:
            try:
                if rows_added == MAX_ROWS_IN_SHEET - 1:
                    logger.debug(
                        "Max row per worksheet limit will be reached. Creating new worksheet..."
                    )
                    self._second_sheet = workbook.create_sheet()
                    self._add_tweet_header(self._second_sheet)

                if rows_added < MAX_ROWS_IN_SHEET:
                    self._sheet.append(FileReporter._get_tweet_row_data(tweet_data_item.data))
                else:
                    self._second_sheet.append(
                        FileReporter._get_tweet_row_data(tweet_data_item.data)
                    )

                rows_added += 1

            except IllegalCharacterError:
                tweet_data_item.data["text"] = ILLEGAL_CHARACTERS_RE.sub(
                    r"", tweet_data_item.data["text"]
                )

                try:
                    if rows_added < MAX_ROWS_IN_SHEET:
                        self._sheet.append(FileReporter._get_tweet_row_data(tweet_data_item.data))
                    else:
                        self._second_sheet.append(
                            FileReporter._get_tweet_row_data(tweet_data_item.data)
                        )

                    rows_added += 1
                except IllegalCharacterError:
                    logger.error(
                        f"Received IllegalCharacterError for id={tweet_data_item.data['id']}. Skipping..."
                    )

        self._adjust_column_widths()

        workbook.save(self._filename)

    def _create_workbook(self) -> openpyxl.Workbook:
        """Create workbook

        :rtype: Workbook
        :returns: Workbook instance
        """

        workbook = openpyxl.Workbook()

        return workbook

    def _add_user_header(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Add user data header

        :type sheet: openpyxl.worksheet.worksheet.Worksheet
        :param sheet: Worksheet to add user data header
        """

        for i, value in enumerate(self._user_data_header, start=1):
            cell = sheet.cell(row=1, column=i)
            cell.value = value
            cell.font = Font(bold=True)

    def _add_tweet_header(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Add tweet data header

        :type sheet: openpyxl.worksheet.worksheet.Worksheet
        :param sheet: Worksheet to add tweet data header
        """

        if self._extracted_data_type == ExtractedDataType.USER_TWEETS:
            self._tweet_data_header.remove("Author")

        for i, value in enumerate(self._tweet_data_header, start=1):
            cell = sheet.cell(row=1, column=i)
            cell.value = value
            cell.font = Font(bold=True)

    def _adjust_column_widths(self) -> None:
        """Adjust the column widths according to maximum length of content"""

        columns = ascii_uppercase[: len(self._user_data_header)]
        sheets = [self._sheet]

        if self._second_sheet:
            sheets.append(self._second_sheet)

        for sheet in sheets:
            column_name = iter(columns)

            for col in sheet.iter_cols():
                contents = []
                for cell in col:
                    contents.append(cell.value)

                max_content_length = max([len(str(content)) for content in contents])
                sheet.column_dimensions[next(column_name)].width = max_content_length + 2
